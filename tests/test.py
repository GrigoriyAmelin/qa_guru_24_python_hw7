import csv
import io

from zipfile import ZipFile
from pypdf import PdfReader
from openpyxl import load_workbook


def test_pdf_file(zip_files):
    with ZipFile(zip_files, mode='r') as zip_archive:
        pdf_bytes_to_read = zip_archive.read('data/pdf_example.pdf')
        pdf_file_to_read = io.BytesIO(pdf_bytes_to_read)
        pdf_reader = PdfReader(pdf_file_to_read)

        print(f'\nPDF file pages counter: {len(pdf_reader.pages)}')
        for page_number in range(len(pdf_reader.pages)):
            page = pdf_reader.pages[page_number]
            text = page.extract_text()
            print(f'PDF file content:\n {(text[:300])}')

        assert len(pdf_reader.pages) is not None
        assert "darkstoreIds" in text
        assert "1fa60713-1cf8-11ea-8d70-0050560306e1" in text


def test_xlsx_file(zip_files):
    with ZipFile(zip_files, mode='r') as zip_archive:
        with zip_archive.open('data/xlsx_example.xlsx') as xlsx_file:
            workbook = load_workbook(xlsx_file)
            sheet = workbook.active
            cell_2_5 = sheet.cell(row=2, column=5).value
            cell_1_2 = sheet.cell(row=1, column=2).value
            cell_2_2 = sheet.cell(row=2, column=2).value

            print('\nXLSX file content:')
            for row in sheet.iter_rows():
                for cell in row:
                    print(cell.value, end='\t')
                print()
            print(f'\nCell 2-5 value: {cell_2_5}')
            print(f'Cell 1-2 value: {cell_1_2}')
            print(f'Cell 2-2 value: {cell_2_2}')

            assert cell_2_5 == 10
            assert cell_1_2 == 'ID склада'
            assert cell_2_2 == '323b9adc-800d-11eb-85a3-1c34dae33151'


def test_csv_file(zip_files):
    with ZipFile(zip_files, mode='r') as zip_archive:
        with zip_archive.open('data/csv_example.csv') as csv_file:
            csv_decoded_content = csv_file.read().decode('utf-8')
            csv_reader = list(csv.reader(csv_decoded_content.splitlines(), delimiter=';'))
            row_one = csv_reader[0]
            row_two = csv_reader[1]
            row_three = csv_reader[2]

            print(f'\nRow number: {len(csv_reader)}')
            for row in range(len(csv_reader)):
                print(f'Row {row}: {csv_reader[row]}')

            assert row_one[2] == 'GUID товара'
            assert row_two[0] == 'тест'
            assert row_three[2] == 'b6fbc6f9-878f-11ed-ae77-08c0eb320147'